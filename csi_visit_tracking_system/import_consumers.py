"""
Import (or re-import) consumers from the agency's CSI/EMedi Stats working
spreadsheet into the SQLite database, classifying each consumer into an
Area (Southers / Southeast / Central) by city, and treating the WSC column
as the assigned Worker.

Safe to re-run any time a new export is dropped in — it upserts by
iConnect Id, so existing visit_records / users history is untouched.

Usage:
    python3 import_consumers.py path/to/"EMedi Stats ... WORKING DOC.xlsx"
"""
import sys
import os
from datetime import datetime

import openpyxl

from db import conn_ctx, init_db
from areas import classify_city


def s(v):
    return str(v).strip() if v is not None else ""


def parse_date(v):
    if not v:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    v = str(v).strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(v, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def import_file(xlsx_path, sheet_name="CSI", uploaded_by=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    init_db()
    imported, skipped = 0, 0
    worker_codes_seen = set()
    area_counts = {}

    with conn_ctx() as conn:
        for r in rows:
            (iid, first, last, _dob, _gender, wsc, address, city, zip_, _email,
             phone, eff, active, _statusdate, hm, il, gh, cdc) = (list(r) + [None] * 18)[:18]

            if iid is None or first is None:
                skipped += 1
                continue
            eff_iso = parse_date(eff)
            if not eff_iso:
                skipped += 1
                continue

            worker_code = s(wsc)
            city_val = s(city)
            area = classify_city(city_val)
            area_counts[area] = area_counts.get(area, 0) + 1
            if worker_code:
                worker_codes_seen.add(worker_code)

            conn.execute(
                """INSERT INTO consumers
                   (id, first_name, last_name, name, worker_code, area, address, city, zip, phone,
                    health_manager, effective_date, active, gh, il, cdc)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                   ON CONFLICT(id) DO UPDATE SET
                     first_name=excluded.first_name, last_name=excluded.last_name,
                     name=excluded.name, worker_code=excluded.worker_code, area=excluded.area,
                     address=excluded.address, city=excluded.city, zip=excluded.zip,
                     phone=excluded.phone, health_manager=excluded.health_manager,
                     effective_date=excluded.effective_date, active=excluded.active,
                     gh=excluded.gh, il=excluded.il, cdc=excluded.cdc""",
                (
                    int(iid), s(first), s(last), f"{s(first)} {s(last)}", worker_code, area,
                    s(address), city_val, s(zip_), s(phone), s(hm), eff_iso,
                    1 if s(active).upper() == "YES" else 0,
                    1 if s(gh).upper() == "YES" else 0,
                    1 if s(il).upper() == "YES" else 0,
                    1 if s(cdc).upper() == "YES" else 0,
                ),
            )
            imported += 1

        for code in sorted(worker_codes_seen):
            conn.execute(
                """INSERT OR IGNORE INTO worker_directory (worker_code, display_name, email, phone)
                   VALUES (?, ?, ?, ?)""",
                (code, code, "", ""),
            )

        conn.execute(
            "INSERT INTO data_uploads (filename, uploaded_at, uploaded_by, row_count) VALUES (?,?,?,?)",
            (os.path.basename(xlsx_path), datetime.utcnow().isoformat(), uploaded_by, imported),
        )

    print(f"Imported {imported} consumers, skipped {skipped} incomplete rows.")
    print(f"Worker codes present: {', '.join(sorted(worker_codes_seen))}")
    print(f"Area breakdown: {area_counts}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 import_consumers.py path/to/spreadsheet.xlsx")
        sys.exit(1)
    import_file(sys.argv[1])
